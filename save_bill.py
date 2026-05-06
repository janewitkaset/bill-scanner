#!/usr/bin/env python3
"""
save_bill.py — รับ JSON จาก stdin แล้วเขียนลง bills.xlsx
ใช้งาน: echo '{"sheet":"florish","data":{...}}' | python3 save_bill.py /path/to/bills.xlsx
"""
import sys
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── ค่าคงที่ ───────────────────────────────────────────────
HEADERS = [
    "timestamp", "vendor_name", "customer_name", "bill_number",
    "bill_date", "amount_before_vat", "vat_amount", "total_amount",
    "confidence",
]
HEADER_LABELS = [
    "วันที่บันทึก", "ชื่อ Vendor", "ชื่อลูกค้า", "เลขที่บิล",
    "วันที่บิล", "ก่อน VAT (฿)", "VAT (฿)", "รวมสุทธิ (฿)",
    "ความแม่นยำ",
]
NUM_COLS = {"amount_before_vat", "vat_amount", "total_amount"}
COL_WIDTHS = [22, 28, 28, 18, 14, 16, 12, 16, 14]

SHEETS = ["florish", "janewit", "unmatched"]

# สีธีม
DARK    = "1A1A2E"
WHITE   = "FFFFFF"
GREEN   = "00E676"
EVEN    = "F0F4FF"
ODD     = "FFFFFF"
BORDER_COLOR = "CBD5E0"


def make_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def setup_sheet(ws):
    """ตั้ง header row สำหรับ Sheet ใหม่"""
    hfill = PatternFill("solid", fgColor=DARK)
    hfont = Font(name="Arial", bold=True, color=WHITE, size=11)
    for ci, label in enumerate(HEADER_LABELS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def find_last_data_row(ws):
    """หาแถวข้อมูลสุดท้าย (ข้าม header และ summary)"""
    last = 1
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val and str(val) not in ("", "รวมทั้งหมด"):
            last = row[0].row
    return last


def write_summary(ws, last_data_row):
    """เขียน/อัปเดตแถว summary ใต้ข้อมูล"""
    summary_row = last_data_row + 2

    # ล้าง summary เก่า (สูงสุด 3 แถว)
    for r in range(last_data_row + 1, last_data_row + 5):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="Arial", size=11)
            cell.border = Border()

    sfill = PatternFill("solid", fgColor=DARK)
    sfont_label = Font(name="Arial", bold=True, color=WHITE, size=11)
    sfont_num   = Font(name="Arial", bold=True, color=GREEN, size=11)

    lbl = ws.cell(row=summary_row, column=1, value="รวมทั้งหมด")
    lbl.fill = sfill
    lbl.font = sfont_label
    lbl.alignment = Alignment(horizontal="center", vertical="center")

    # SUM สำหรับ amount_before_vat(col6), vat_amount(col7), total_amount(col8)
    for ci in [6, 7, 8]:
        col_letter = get_column_letter(ci)
        c = ws.cell(
            row=summary_row, column=ci,
            value=f"=SUM({col_letter}2:{col_letter}{last_data_row})"
        )
        c.fill = sfill
        c.font = sfont_num
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")


def append_row(ws, row_data):
    """เพิ่มแถวข้อมูล 1 แถว"""
    last = find_last_data_row(ws)
    new_row = last + 1

    fill = PatternFill("solid", fgColor=EVEN if new_row % 2 == 0 else ODD)
    border = make_border()

    for ci, key in enumerate(HEADERS, 1):
        val = row_data.get(key, "")
        cell = ws.cell(row=new_row, column=ci, value=val)
        cell.fill = fill
        cell.border = border
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if key in NUM_COLS:
            try:
                cell.value = float(val) if val else 0
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            except (ValueError, TypeError):
                pass

    write_summary(ws, new_row)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 save_bill.py /path/to/bills.xlsx", file=sys.stderr)
        sys.exit(1)

    fpath = Path(sys.argv[1])
    payload = json.loads(sys.stdin.read())
    sheet_name = payload["sheet"]   # "florish" | "janewit" | "unmatched"
    row_data   = payload["data"]

    # โหลดหรือสร้างไฟล์ใหม่
    if fpath.exists():
        wb = load_workbook(fpath)
    else:
        wb = Workbook()
        wb.remove(wb.active)           # ลบ Sheet เริ่มต้น
        for s in SHEETS:
            ws = wb.create_sheet(s)
            setup_sheet(ws)

    # สร้าง Sheet ถ้ายังไม่มี
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        setup_sheet(ws)

    append_row(wb[sheet_name], row_data)
    wb.save(fpath)
    print("ok")


if __name__ == "__main__":
    main()
